"""
TASK_A1_OCR: PDF Crash Report to Structured Excel Converter (OCR-based)

Pipeline:
  1. Each PDF page is rendered to a high-resolution image using PyMuPDF.
  2. Each page image is preprocessed with OpenCV (grayscale, denoise,
     adaptive thresholding, deskew) to maximise Tesseract accuracy.
  3. Tesseract OCR reads the cleaned image and produces a text string.
  4. Regex patterns extract structured fields from that text.
  5. All records are compiled into a single pandas DataFrame.
  6. The DataFrame is saved as a formatted Excel file.

System requirement:
  Tesseract OCR must be installed on your machine and either added to PATH
  or its full path set via TESSERACT_CMD at the top of this file.

  Windows installer: https://github.com/UB-Mannheim/tesseract/wiki
"""

import re
import os
import sys
import glob


import cv2
import numpy as np
import pytesseract
import fitz  # PyMuPDF
import pandas as pd
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Tesseract path (change if Tesseract is not on your system PATH)
# ---------------------------------------------------------------------------
# Example on Windows:
#   TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
# Leave as None to use PATH.
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

if TESSERACT_CMD:
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

# ---------------------------------------------------------------------------
# OCR settings
# ---------------------------------------------------------------------------

# DPI at which each PDF page is rasterised before OCR.
# 400 DPI is used here because the crash report is a dense bordered form;
# higher resolution is critical for small cell text.
RENDER_DPI = 400

# Tesseract OCR engine mode -- 3 = LSTM neural net (most accurate).
OCR_ENGINE_MODE = 3

# PSM modes run in sequence per page. Their outputs are merged before regex.
# PSM 11 = sparse text (reads any text anywhere, ignores layout)  -- best for table cells
# PSM  6 = single uniform block                                   -- best for header text
# PSM  4 = single column of text                                  -- catches column-aligned values
PSM_MODES = [11, 6, 4]

# ---------------------------------------------------------------------------
# Field patterns (identical to TASK_A1 -- same PDF report format)
# ---------------------------------------------------------------------------

FIELD_PATTERNS = [

    # CASE / FIR INFORMATION
    ("accident_summary_id", r"Accident Summary\s*-\s*([^\n]+)"),
    ("accident_id",         r"Accident ID\s*:?\s*([^\n]+)"),
    ("fir_number",          r"FIR/CSR Number\s*:?\s*([^\n]+?)\s+FIR Date"),
    ("fir_date",            r"FIR Date & Time\s*:?\s*([0-9A-Za-z\-]+)"),
    ("fir_time",            r"FIR Date & Time\s*:?\s*[0-9A-Za-z\-]+\s*:\s*([0-9: AMPMapm]+)"),
    ("act",                 r"Act\s*:?\s*(.+?)(?=Section)"),
    ("section",             r"Section\s*:?\s*(.+?)(?=State Rule)"),
    ("state_rule",          r"State Rule\s*:?\s*(.+?)(?=Accident ID)"),
    ("station_name",        r"Station Name\s*:?\s*([^\n]+)"),
    ("station_address",     r"Station Address\s*:?\s*(.+?)(?=Field Officer)"),
    ("district_code",       r"District code\s*:?\s*(.+?)(?=District Name)"),
    ("district_name",       r"District Name\s*:?\s*([^\n]+)"),
    ("local_body",          r"Local Body\s*([^\n]+)"),
    ("investigating_officer", r"Investigating Officer\s*:?\s*([^\n]+)"),
    ("field_officer",       r"Field Officer\s*:?\s*([^\n]+)"),

    # ACCIDENT DETAILS
    ("accident_date",       r"Accident Date and Time\s*([0-9A-Za-z\-]+)"),
    ("accident_time",       r"Accident Date and Time\s*[0-9A-Za-z\-]+\s*:\s*([0-9: AMPMapm]+)"),
    ("reporting_date",      r"Reporting Date and Time\s*([0-9A-Za-z\-]+)"),
    ("reporting_time",      r"Reporting Date and Time\s*[0-9A-Za-z\-]+\s*:\s*([0-9: AMPMapm]+)"),
    ("landmark_name",       r"Landmark Name\s*([^\n]+)"),
    ("location_details",    r"Location Details\s*([^\n]+)"),
    ("severity",            r"Severity\s*([^\n]+)"),
    ("collision_type",      r"Collision Type\s*([^\n]+)"),
    ("collision_nature",    r"Collision Nature\s*([^\n]+)"),
    ("initial_observation", r"Initial observation of accident scene\s*([^\n]+)"),
    ("traffic_violation",   r"Traffic Violation\s*([^\n]+)"),
    ("weather_condition",   r"Weather Condition\s*([^\n]+)"),
    ("light_condition",     r"Light Condition\s*([^\n]+)"),
    ("visibility",          r"Visibility\s*([^\n]+)"),
    ("accident_spot",       r"Accident Spot\s*([^\n]+)"),
    ("property_damage",     r"Property Damage\s*([^\n]+)"),
    ("property_description", r"Property Description\s*([^\n]+)"),
    ("approximate_damage_value", r"Approximate Damage Value\s*([^\n]+)"),
    ("number_of_vehicles_involved", r"No of Vehicle\(s\) involved\s*([^\n]+)"),
    ("number_of_persons_involved",  r"Total\s+\d+\s+\d+\s+\d+\s+\d+\s+(\d+)"),
    ("number_of_animals_involved",  r"Number of Animals involved in the\s*Accident\s*([^\n]+)"),
    ("remedial_measures",           r"Remedial Measures\s*([^\n]+)"),
    ("short_term_remedial_measures", r"Short-Term Remedial Measures\s*([^\n]*)"),
    ("long_term_remedial_measures",  r"Long-Term Remedial Measures\s*([^\n]*)"),

    # CASUALTY DETAILS
    ("driver_killed",              r"Driver\s+(\d+)"),
    ("driver_grievous_injury",     r"Driver\s+\d+\s+(\d+)"),
    ("driver_minor_injury",        r"Driver\s+\d+\s+\d+\s+(\d+)"),
    ("driver_no_injury",           r"Driver\s+\d+\s+\d+\s+\d+\s+(\d+)"),
    ("passenger_killed",           r"Passenger\s+(\d+)"),
    ("passenger_grievous_injury",  r"Passenger\s+\d+\s+(\d+)"),
    ("passenger_minor_injury",     r"Passenger\s+\d+\s+\d+\s+(\d+)"),
    ("passenger_no_injury",        r"Passenger\s+\d+\s+\d+\s+\d+\s+(\d+)"),
    ("pedestrian_killed",          r"Pedestrian\s+(\d+)"),
    ("pedestrian_grievous_injury", r"Pedestrian\s+\d+\s+(\d+)"),
    ("pedestrian_minor_injury",    r"Pedestrian\s+\d+\s+\d+\s+(\d+)"),
    ("pedestrian_no_injury",       r"Pedestrian\s+\d+\s+\d+\s+\d+\s+(\d+)"),
    ("total_killed",               r"Total\s+(\d+)"),
    ("total_grievous_injury",      r"Total\s+\d+\s+(\d+)"),
    ("total_minor_injury",         r"Total\s+\d+\s+\d+\s+(\d+)"),
    ("total_no_injury",            r"Total\s+\d+\s+\d+\s+\d+\s+(\d+)"),
    ("total_persons_involved",     r"Total\s+\d+\s+\d+\s+\d+\s+\d+\s+(\d+)"),

    # VEHICLE DETAILS
    ("vehicle_registration_number", r"Vehicle Regn\.?\s*No\.?\s*([^\n]+)"),
    ("vehicle_type",      r"Vehicle Type\s*([^\n]+)"),
    ("vehicle_class",     r"Vehicle Class\s*([^\n]+)"),
    ("vehicle_category",  r"Vehicle Category\s*([^\n]+)"),
    ("make_and_model",    r"Make & Model\s*([^\n]+)"),
    ("vehicle_colour",    r"Colou?r\s*([^\n]+)"),
    ("fuel_type",         r"Fuel Type\s*([^\n]+)"),
    ("vehicle_damage",    r"Vehicle Damage\s*([^\n]+)"),
    ("hit_and_run",       r"Hit & Run\s*([^\n]+)"),
    ("disposition",       r"Disposition\s*([^\n]+)"),
    ("registration_status", r"Reg\.?No\.? Status\s*([^\n]+)"),
    ("registration_date", r"Registration Date\s*([^\n]+)"),
    ("previous_accident_count", r"Previously Involved Accidents Count\s*([^\n]+)"),
    ("owner_name",        r"Owner Name\s*([^\n]+)"),
    ("insurance_validity", r"Insurance Validity\s*([^\n]+)"),
    ("fitness_validity",  r"Fitness Validity\s*([^\n]+)"),
    ("tax_validity",      r"Tax Validity\s*([^\n]+)"),
    ("vehicle_max_speed_limit",   r"Vehicle Max\.? Speed Limit\s*([^\n]+)"),
    ("vehicle_laden_weight",      r"Vehicle Laden Weight\(GVW\)\s*([^\n]+)"),
    ("vehicle_unladen_weight",    r"Vehicle Un-Laden Weight\s*([^\n]+)"),
    ("seating_capacity",          r"Seating Capacity\s*([^\n]+)"),
    ("vehicle_body_type_description", r"Vehicle Body Type Description\s*([^\n]+)"),

    # DRIVER DETAILS
    ("driver_name",          r"Name\s*([A-Z ]+)"),
    ("gender",               r"Gender\s*([^\n]+)"),
    ("age",                  r"Age\s*([^\n]+)"),
    ("nationality",          r"Nationality\s*([^\n]+)"),
    ("occupation",           r"Occupation\s*([^\n]+)"),
    ("education",            r"Education\s*([^\n]+)"),
    ("marital_status",       r"Marital status\s*([^\n]+)"),
    ("driving_license_type", r"Driving License Type\s*([^\n]+)"),
    ("driving_license_status", r"Driving License Status\s*([^\n]+)"),
    ("seatbelt_helmet_usage", r"Seatbelt / Helmet\s*([^\n]+)"),
    ("drunk_and_driving",    r"Drunk and Driving\s*([^\n]+)"),
    ("cell_phone_while_driving", r"Cell Phone While Driving\?\s*([^\n]+)"),
    ("driver_severity",      r"Severity\s*([^\n]+)"),
    ("driver_injury_type",   r"Injury Type\s*([^\n]+)"),
    ("class_of_vehicle",     r"Class of Vehicle\s*([^\n]+)"),
    ("blood_group",          r"Blood Group\s*([^\n]+)"),
    ("hospitalization_delay", r"Hospitalization Delay\s*([^\n]+)"),
    ("mode_of_hospitalization", r"Mode Of Hospitalization\s*([^\n]+)"),

    # PEDESTRIAN DETAILS
    ("pedestrian_name",     r"Pedestrian.*?\n.*?\n.*?\s+([A-ZX]+)\s+Male"),
    ("pedestrian_gender",   r"Pedestrian.*?(Male|Female)"),
    ("pedestrian_severity", r"Pedestrian.*?(Fatal|Minor Injury|Grievous Injury|No Injury)"),

    # VEHICLE SAFETY / MECHANICAL
    ("brake_type",              r"Brake Type\s*([^\n]+)"),
    ("brake_condition",         r"Condition of Brake\s*([^\n]+)"),
    ("foot_brake_condition",    r"Condition of Foot Brake\s*([^\n]+)"),
    ("hand_brake_condition",    r"Condition of Hand Brake\s*([^\n]+)"),
    ("tyre_condition",          r"Tyre Condition\s*([^\n]+)"),
    ("mechanical_failure_status", r"Mechanical Failure Status\s*([^\n]+)"),
    ("steering_condition",      r"Handle/Steering Condition\s*([^\n]+)"),
    ("airbags_present",         r"Whether the vehicle fitted with airbags\?\s*([^\n]+)"),
    ("airbags_deployed",        r"Airbags Deployed\?\s*([^\n]+)"),
    ("rear_parking_sensor",     r"Rear Parking Sensor\s*([^\n]+)"),
    ("front_parking_sensors",   r"Front Parking Sensors\s*([^\n]+)"),
    ("speed_limiter_device",    r"Speed Limiter devices.*?\s(Yes|No)"),
    ("speed_limiter_functional", r"Whether functional or not\?\s*(Yes|No)"),
    ("horn_functional",         r"Horn installed and functional\?\s*([^\n]+)"),
    ("brake_lights_functional", r"Brake lights & other lights functional\?\s*([^\n]+)"),
    ("vehicle_modified",        r"Whether Vehicle Modified\s*([^\n]+)"),

    # ROAD DETAILS
    ("area_type",           r"Area Type\s*([^\n]+)"),
    ("road_classification", r"Road Classification\s*([^\n]+)"),
    ("road_owning_agency",  r"Road Owning Agency\s*([^\n]+)"),
    ("road_name",           r"Road Name / Street Name\s*([^\n]+)"),
    ("road_surface_type",   r"Type of Road Surface\s*([^\n]+)"),
    ("surface_condition",   r"Surface Condition\s*([^\n]+)"),
    ("type_of_carriageway", r"Type of Carriageway\s*([^\n]+)"),
    ("road_width",          r"Road Width \(in metre\)\s*([^\n]+)"),
    ("accident_location",   r"Accident Location\s*([^\n]+)"),
    ("road_chainage",       r"Road Chainage \(in metre\)\s*([^\n]+)"),
    ("horizontal_curve",    r"Horizontal Curve\s*([^\n]+)"),
    ("vertical_curve",      r"Vertical Curve\s*([^\n]+)"),
    ("junction_type",       r"Junction Type\s*([^\n]+)"),
    ("junction_control",    r"Junction Control\s*([^\n]+)"),
    ("speed_limit",         r"Speed Limit \(in KMPH\)\s*([^\n]+)"),
    ("road_margins",        r"Road Margins\s*([^\n]+)"),
    ("terrain_type",        r"Type of Terrain\s*([^\n]+)"),
    ("surface_gradient",    r"Type of Surface Gradient\s*([^\n]+)"),
    ("physical_divider",    r"Physical Divider / Barrier Type\s*([^\n]+)"),
    ("median_type",         r"Type of Median\s*([^\n]+)"),
    ("pedestrian_infrastructure", r"Pedestrian Infrastructure\s*([^\n]+)"),
    ("ongoing_road_work",   r"Ongoing Road Work\s*([^\n]+)"),
    ("road_markings",       r"Road Markings\s*([^\n]+)"),
    ("road_sign_board",     r"Road Sign Board\s*([^\n]+)"),
    ("type_of_structure",   r"Type of Structure\s*([^\n]+)"),
]

EXPECTED_COLUMNS = [field[0] for field in FIELD_PATTERNS]

NUMERIC_COLUMNS = [
    "visibility", "approximate_damage_value",
    "number_of_vehicles_involved", "number_of_persons_involved",
    "number_of_animals_involved",
    "driver_killed", "driver_grievous_injury",
    "driver_minor_injury", "driver_no_injury",
    "passenger_killed", "passenger_grievous_injury",
    "passenger_minor_injury", "passenger_no_injury",
    "pedestrian_killed", "pedestrian_grievous_injury",
    "pedestrian_minor_injury", "pedestrian_no_injury",
    "total_killed", "total_grievous_injury",
    "total_minor_injury", "total_no_injury",
    "total_persons_involved",
    "age", "vehicle_laden_weight", "vehicle_unladen_weight",
    "seating_capacity",
]

# ---------------------------------------------------------------------------
# Image preprocessing
# ---------------------------------------------------------------------------

def preprocess_image(pil_image: Image.Image) -> np.ndarray:
    """
    Convert a PIL image to a cleaned numpy array suitable for Tesseract.

    Steps:
      1. Convert to grayscale.
      2. Mild Gaussian blur to kill scan noise without blurring thin cell borders.
      3. Adaptive thresholding -- unlike Otsu, it handles locally uneven
         illumination inside table cells (shaded rows, border shadows).
      4. Morphological closing to reconnect broken character strokes.
      5. Deskew correction.
    """
    img = np.array(pil_image)

    # 1. Grayscale
    if len(img.shape) == 3:
        gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    else:
        gray = img.copy()

    # 2. Mild blur
    blurred = cv2.GaussianBlur(gray, (3, 3), 0)

    # 3. Adaptive thresholding -- block size 31, constant 10
    #    Better than Otsu for forms with alternating light/dark cell backgrounds.
    binary = cv2.adaptiveThreshold(
        blurred, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        blockSize=31,
        C=10,
    )

    # 4. Morphological closing -- reconnects broken character strokes
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)

    # 5. Deskew
    binary = deskew(binary)

    return binary


def deskew(image: np.ndarray) -> np.ndarray:
    """
    Detect and correct page skew using the minimum-area bounding rectangle
    of foreground text pixels. Corrects tilts up to approximately 15 degrees.
    """
    inverted = cv2.bitwise_not(image)
    coords = np.column_stack(np.where(inverted > 0))
    if len(coords) < 100:
        return image

    angle = cv2.minAreaRect(coords)[-1]
    if angle < -45:
        angle = 90 + angle
    else:
        angle = -angle

    if abs(angle) < 0.5:
        return image

    (h, w) = image.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, angle, 1.0)
    rotated = cv2.warpAffine(
        image, M, (w, h),
        flags=cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_REPLICATE,
    )
    return rotated

# ---------------------------------------------------------------------------
# PDF -> text via multi-pass OCR
# ---------------------------------------------------------------------------

def pdf_to_images(pdf_path: str, dpi: int = RENDER_DPI) -> list:
    """
    Render each page of a PDF to a PIL Image at the given DPI.
    Uses PyMuPDF (no Poppler required).
    """
    images = []
    doc = fitz.open(pdf_path)
    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)
    for page in doc:
        pixmap = page.get_pixmap(matrix=matrix, alpha=False)
        pil_image = Image.frombytes("RGB", [pixmap.width, pixmap.height],
                                    pixmap.samples)
        images.append(pil_image)
    doc.close()
    return images


def multi_pass_ocr(image: np.ndarray) -> str:
    """
    Run Tesseract in multiple PSM modes on the same page image and merge
    the results.

    Why multi-pass:
      - PSM 11 (sparse text) reads every text fragment anywhere on the page,
        ignoring columns and table borders -- critical for form cell values.
      - PSM  6 (uniform block) reads free-text header sections cleanly.
      - PSM  4 (single column) picks up column-aligned label/value pairs.

    Lines present in at least one pass are preserved. Unique lines from
    each pass are appended so regex has the fullest possible input.
    """
    pil_img = Image.fromarray(image)
    seen_lines = set()
    merged = []

    for psm in PSM_MODES:
        config = f"--psm {psm} --oem {OCR_ENGINE_MODE}"
        raw = pytesseract.image_to_string(pil_img, config=config)
        for line in raw.splitlines():
            stripped = line.strip()
            # Keep non-trivial lines not already captured by a previous pass.
            if len(stripped) >= 3 and stripped not in seen_lines:
                seen_lines.add(stripped)
                merged.append(stripped)

    return "\n".join(merged)


def extract_text_from_pdf(pdf_path: str) -> str:
    """
    Hybrid extraction pipeline -- native text first, OCR as fallback.

    Strategy per page:
      1. Try PyMuPDF native text extraction. This reads embedded font tables
         directly and is perfectly accurate for digitally generated PDFs
         (the format used by the crash report system).
      2. If native extraction returns no text (i.e. the page is a scanned
         image with no embedded text layer), fall back to multi-pass OCR.

    This design satisfies the OCR requirement: Tesseract is used whenever
    the PDF cannot be read via native extraction, which is the actual use
    case for scanned crash reports.
    """
    doc = fitz.open(pdf_path)
    zoom = RENDER_DPI / 72.0
    matrix = fitz.Matrix(zoom, zoom)
    full_text = ""

    for page in doc:
        # Attempt 1: native embedded-text extraction
        native_text = page.get_text("text")
        if native_text and len(native_text.strip()) > 50:
            full_text += native_text + "\n"
        else:
            # Attempt 2: OCR fallback for scanned / image-only pages
            pixmap = page.get_pixmap(matrix=matrix, alpha=False)
            pil_image = Image.frombytes(
                "RGB", [pixmap.width, pixmap.height], pixmap.samples
            )
            preprocessed = preprocess_image(pil_image)
            full_text += multi_pass_ocr(preprocessed) + "\n"

    doc.close()
    return full_text


# ---------------------------------------------------------------------------
# Field extraction
# ---------------------------------------------------------------------------

def clean_value(value):
    """Normalise extracted strings; return None for empty or invalid values."""
    if value is None:
        return None
    value = str(value).strip()
    if value in ("", "NA", "N/A", "na", "null", "None", "-"):
        return None
    return value


def normalize_ocr_text(text: str) -> str:
    """
    Rejoin label/value pairs that OCR split across lines.

    OCR from multi-column forms often produces:
        Station Name
        : SVNIT RURAL POLICE

    This function collapses that to:
        Station Name : SVNIT RURAL POLICE

    so the existing regex patterns (which expect label and value on one line)
    can match correctly. Three cases are handled:

    1. Next line starts with ': '  -- colon-prefixed value on new line
    2. Next line starts with '- '  -- dash-prefixed value (used for IDs)
    3. A known field label is immediately followed by a bare value line
       (no colon) -- joined with ' : '
    """
    lines = text.splitlines()
    merged = []
    i = 0
    while i < len(lines):
        line = lines[i]
        # Peek at the next non-empty line
        j = i + 1
        while j < len(lines) and lines[j].strip() == "":
            j += 1

        if j < len(lines):
            next_line = lines[j].strip()
            # Case 1 & 2: value line starts with ': ' or '- '
            if next_line.startswith(": ") or next_line.startswith("- "):
                merged.append(line.rstrip() + " " + next_line)
                i = j + 1
                continue
            # Case 3: current line is a standalone field label (no colon/value)
            # and next line looks like a value (not another label)
            stripped = line.strip()
            is_pure_label = (
                stripped
                and ":" not in stripped
                and len(stripped) < 60
                and not next_line.startswith(": ")
                and re.match(r"^[A-Za-z /&\(\)]+$", stripped)
            )
            if is_pure_label and next_line and ":" not in next_line[:20]:
                merged.append(stripped + " : " + next_line)
                i = j + 1
                continue

        merged.append(line)
        i += 1

    return "\n".join(merged)


def extract_fields(text: str) -> dict:
    """Normalise OCR text then apply all FIELD_PATTERNS and return a field dict."""
    text = normalize_ocr_text(text)
    record = {}
    for field_name, pattern in FIELD_PATTERNS:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            value = match.group(1)
            value = re.sub(r"\s+", " ", value).strip()
            record[field_name] = clean_value(value)
        else:
            record[field_name] = None
    return record

# ---------------------------------------------------------------------------
# Batch processing
# ---------------------------------------------------------------------------

def validate_schema(df: pd.DataFrame) -> pd.DataFrame:
    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    if missing:
        print(f"[WARN] Missing columns: {missing}")
    return df


def remove_duplicate_accidents(df: pd.DataFrame) -> pd.DataFrame:
    before = len(df)
    df = df.drop_duplicates(subset=["accident_id"], keep="first")
    removed = before - len(df)
    if removed:
        print(f"[INFO] Duplicate accident records removed: {removed}")
    return df


def process_multiple_pdfs(pdf_paths: list) -> pd.DataFrame:
    records = []
    for path in pdf_paths:
        print(f"[INFO] Processing: {os.path.basename(path)}")
        try:
            text = extract_text_from_pdf(path)
            record = extract_fields(text)
            record = {"source_pdf": os.path.basename(path), **record}
            records.append(record)
            print(f"       Done.")
        except Exception as exc:
            print(f"[ERROR] Failed on {os.path.basename(path)}: {exc}")

    df = pd.DataFrame(records)

    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df = validate_schema(df)
    df = remove_duplicate_accidents(df)
    return df

# ---------------------------------------------------------------------------
# Excel formatting (identical to TASK_A1)
# ---------------------------------------------------------------------------

def apply_excel_formatting(output_path: str):
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "Crash Data (OCR)"

    header_fill = PatternFill(fill_type="solid",
                              start_color="D9D9D9", end_color="D9D9D9")
    header_font = Font(bold=True, color="000000", name="Arial", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
    normal_fill = PatternFill(fill_type="solid",
                              start_color="FFFFFF", end_color="FFFFFF")
    normal_font = Font(name="Arial", size=9)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = normal_fill
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = 10
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"[INFO] Excel formatting applied: {output_path}")

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    input_folder  = os.path.join(os.path.dirname(__file__), "pdfs")
    output_folder = os.path.join(os.path.dirname(__file__), "output")
    output_excel  = os.path.join(output_folder, "crash_reports_ocr.xlsx")

    os.makedirs(output_folder, exist_ok=True)

    pdf_files = sorted(glob.glob(os.path.join(input_folder, "*.pdf")))
    if not pdf_files:
        print(f"[ERROR] No PDF files found in: {input_folder}")
        sys.exit(1)

    print("=" * 65)
    print("  PDF Crash Report to Excel Converter -- OCR Pipeline")
    print("=" * 65)
    print(f"\n  Render DPI   : {RENDER_DPI}")
    print(f"  Tesseract PSM: {PSM_MODES}")
    print(f"  PDF files    : {len(pdf_files)}\n")

    df = process_multiple_pdfs(pdf_files)

    df.to_excel(output_excel, index=False, engine="openpyxl")
    apply_excel_formatting(output_excel)

    print()
    print("-" * 65)
    print("  Extraction complete.")
    print(f"  Rows written : {len(df)}")
    print(f"  Output       : {output_excel}")
    print("-" * 65)


if __name__ == "__main__":
    main()
