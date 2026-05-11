"""
Task A1: PDF Crash Report to Structured Excel Converter
"""

import re
import os
import sys
import glob
import pdfplumber
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FIELD_PATTERNS = [

    # =========================================================
    # CASE / FIR INFORMATION
    # =========================================================

    ("accident_summary_id", r"Accident Summary\s*-\s*([^\n]+)"),
    ("accident_id", r"Accident ID\s*:?\s*([^\n]+)"),
    ("fir_number", r"FIR/CSR Number\s*:?\s*([^\n]+?)\s+FIR Date"),
    ("fir_date", r"FIR Date & Time\s*:?\s*([0-9A-Za-z\-]+)"),
    ("fir_time", r"FIR Date & Time\s*:?\s*[0-9A-Za-z\-]+\s*:\s*([0-9: AMPMapm]+)"),

    ("act",r"Act\s*:?\s*(.+?)(?=Section)"),
    ("section", r"Section\s*:?\s*(.+?)(?=State Rule)"),
    ("state_rule",r"State Rule\s*:?\s*(.+?)(?=Accident ID)"),

    ("station_name", r"Station Name\s*:?\s*([^\n]+)"),
    ("station_address",r"Station Address\s*:?\s*(.+?)(?=Field Officer)"),

    ("district_code",r"District code\s*:?\s*(.+?)(?=District Name)"),
    ("district_name", r"District Name\s*:?\s*([^\n]+)"),

    ("local_body", r"Local Body\s*([^\n]+)"),

    ("investigating_officer", r"Investigating Officer\s*:?\s*([^\n]+)"),
    ("field_officer", r"Field Officer\s*:?\s*([^\n]+)"),

    # =========================================================
    # ACCIDENT DETAILS
    # =========================================================

    ("accident_date", r"Accident Date and Time\s*([0-9A-Za-z\-]+)"),
    ("accident_time", r"Accident Date and Time\s*[0-9A-Za-z\-]+\s*:\s*([0-9: AMPMapm]+)"),

    ("reporting_date", r"Reporting Date and Time\s*([0-9A-Za-z\-]+)"),
    ("reporting_time", r"Reporting Date and Time\s*[0-9A-Za-z\-]+\s*:\s*([0-9: AMPMapm]+)"),

    ("landmark_name", r"Landmark Name\s*([^\n]+)"),
    ("location_details", r"Location Details\s*([^\n]+)"),

    ("severity", r"Severity\s*([^\n]+)"),

    ("collision_type", r"Collision Type\s*([^\n]+)"),
    ("collision_nature", r"Collision Nature\s*([^\n]+)"),

    ("initial_observation", r"Initial observation of accident scene\s*([^\n]+)"),
    ("traffic_violation", r"Traffic Violation\s*([^\n]+)"),

    ("weather_condition", r"Weather Condition\s*([^\n]+)"),
    ("light_condition", r"Light Condition\s*([^\n]+)"),

    ("visibility", r"Visibility\s*([^\n]+)"),

    ("accident_spot", r"Accident Spot\s*([^\n]+)"),

    ("property_damage", r"Property Damage\s*([^\n]+)"),
    ("property_description", r"Property Description\s*([^\n]+)"),
    ("approximate_damage_value", r"Approximate Damage Value\s*([^\n]+)"),

    ("number_of_vehicles_involved", r"No of Vehicle\(s\) involved\s*([^\n]+)"),

    ("number_of_persons_involved",
     r"Total\s+\d+\s+\d+\s+\d+\s+\d+\s+(\d+)"),

    ("number_of_animals_involved",
     r"Number of Animals involved in the\s*Accident\s*([^\n]+)"),

    ("remedial_measures", r"Remedial Measures\s*([^\n]+)"),
    ("short_term_remedial_measures", r"Short-Term Remedial Measures\s*([^\n]*)"),
    ("long_term_remedial_measures", r"Long-Term Remedial Measures\s*([^\n]*)"),

    # =========================================================
    # CASUALTY DETAILS
    # =========================================================

    ("driver_killed", r"Driver\s+(\d+)"),
    ("driver_grievous_injury", r"Driver\s+\d+\s+(\d+)"),
    ("driver_minor_injury", r"Driver\s+\d+\s+\d+\s+(\d+)"),
    ("driver_no_injury", r"Driver\s+\d+\s+\d+\s+\d+\s+(\d+)"),

    ("passenger_killed", r"Passenger\s+(\d+)"),
    ("passenger_grievous_injury", r"Passenger\s+\d+\s+(\d+)"),
    ("passenger_minor_injury", r"Passenger\s+\d+\s+\d+\s+(\d+)"),
    ("passenger_no_injury", r"Passenger\s+\d+\s+\d+\s+\d+\s+(\d+)"),

    ("pedestrian_killed", r"Pedestrian\s+(\d+)"),
    ("pedestrian_grievous_injury", r"Pedestrian\s+\d+\s+(\d+)"),
    ("pedestrian_minor_injury", r"Pedestrian\s+\d+\s+\d+\s+(\d+)"),
    ("pedestrian_no_injury", r"Pedestrian\s+\d+\s+\d+\s+\d+\s+(\d+)"),

    ("total_killed", r"Total\s+(\d+)"),
    ("total_grievous_injury", r"Total\s+\d+\s+(\d+)"),
    ("total_minor_injury", r"Total\s+\d+\s+\d+\s+(\d+)"),
    ("total_no_injury", r"Total\s+\d+\s+\d+\s+\d+\s+(\d+)"),
    ("total_persons_involved", r"Total\s+\d+\s+\d+\s+\d+\s+\d+\s+(\d+)"),

    # =========================================================
    # VEHICLE DETAILS
    # =========================================================

    ("vehicle_registration_number", r"Vehicle Regn\.\s*No\s*([^\n]+)"),
    ("vehicle_type", r"Vehicle Type\s*([^\n]+)"),
    ("vehicle_class", r"Vehicle Class\s*([^\n]+)"),
    ("vehicle_category", r"Vehicle Category\s*([^\n]+)"),

    ("make_and_model", r"Make & Model\s*([^\n]+)"),
    ("vehicle_colour", r"Colour\s*([^\n]+)"),
    ("fuel_type", r"Fuel Type\s*([^\n]+)"),

    ("vehicle_damage", r"Vehicle Damage\s*([^\n]+)"),

    ("hit_and_run", r"Hit & Run\s*([^\n]+)"),
    ("disposition", r"Disposition\s*([^\n]+)"),

    ("registration_status", r"Reg\.No Status\s*([^\n]+)"),
    ("registration_date", r"Registration Date\s*([^\n]+)"),

    ("previous_accident_count",
     r"Previously Involved Accidents Count\s*([^\n]+)"),

    ("owner_name", r"Owner Name\s*([^\n]+)"),

    ("insurance_validity", r"Insurance Validity\s*([^\n]+)"),
    ("fitness_validity", r"Fitness Validity\s*([^\n]+)"),
    ("tax_validity", r"Tax Validity\s*([^\n]+)"),

    ("vehicle_max_speed_limit",
     r"Vehicle Max\. Speed Limit\s*([^\n]+)"),

    ("vehicle_laden_weight",
     r"Vehicle Laden Weight\(GVW\)\s*([^\n]+)"),

    ("vehicle_unladen_weight",
     r"Vehicle Un-Laden Weight\s*([^\n]+)"),

    ("seating_capacity", r"Seating Capacity\s*([^\n]+)"),

    ("vehicle_body_type_description",
     r"Vehicle Body Type Description\s*([^\n]+)"),

    # =========================================================
    # DRIVER DETAILS
    # =========================================================

    ("driver_name", r"Name\s*([A-Z ]+)"),
    ("gender", r"Gender\s*([^\n]+)"),
    ("age", r"Age\s*([^\n]+)"),
    ("nationality", r"Nationality\s*([^\n]+)"),

    ("occupation", r"Occupation\s*([^\n]+)"),
    ("education", r"Education\s*([^\n]+)"),
    ("marital_status", r"Marital status\s*([^\n]+)"),

    ("driving_license_type", r"Driving License Type\s*([^\n]+)"),
    ("driving_license_status", r"Driving License Status\s*([^\n]+)"),

    ("seatbelt_helmet_usage", r"Seatbelt / Helmet\s*([^\n]+)"),

    ("drunk_and_driving", r"Drunk and Driving\s*([^\n]+)"),
    ("cell_phone_while_driving",
     r"Cell Phone While Driving\?\s*([^\n]+)"),

    ("driver_severity", r"Severity\s*([^\n]+)"),
    ("driver_injury_type", r"Injury Type\s*([^\n]+)"),

    ("class_of_vehicle", r"Class of Vehicle\s*([^\n]+)"),

    ("blood_group", r"Blood Group\s*([^\n]+)"),

    ("hospitalization_delay",
     r"Hospitalization Delay\s*([^\n]+)"),

    ("mode_of_hospitalization",
     r"Mode Of Hospitalization\s*([^\n]+)"),

    # =========================================================
    # PEDESTRIAN DETAILS
    # =========================================================

    ("pedestrian_name",
     r"Pedestrian.*?\n.*?\n.*?\s+([A-ZX]+)\s+Male"),

    ("pedestrian_gender",
     r"Pedestrian.*?(Male|Female)"),

    ("pedestrian_severity",
     r"Pedestrian.*?(Fatal|Minor Injury|Grievous Injury|No Injury)"),

    # =========================================================
    # VEHICLE SAFETY / MECHANICAL
    # =========================================================

    ("brake_type", r"Brake Type\s*([^\n]+)"),
    ("brake_condition", r"Condition of Brake\s*([^\n]+)"),

    ("foot_brake_condition",
     r"Condition of Foot Brake\s*([^\n]+)"),

    ("hand_brake_condition",
     r"Condition of Hand Brake\s*([^\n]+)"),

    ("tyre_condition", r"Tyre Condition\s*([^\n]+)"),

    ("mechanical_failure_status",
     r"Mechanical Failure Status\s*([^\n]+)"),

    ("steering_condition",
     r"Handle/Steering Condition\s*([^\n]+)"),

    ("airbags_present",
     r"Whether the vehicle fitted with airbags\?\s*([^\n]+)"),

    ("airbags_deployed",
     r"Airbags Deployed\?\s*([^\n]+)"),

    ("rear_parking_sensor",
     r"Rear Parking Sensor\s*([^\n]+)"),

    ("front_parking_sensors",
     r"Front Parking Sensors\s*([^\n]+)"),

    ("speed_limiter_device",
     r"Speed Limiter devices.*?\s(Yes|No)"),

    ("speed_limiter_functional",
     r"Whether functional or not\?\s*(Yes|No)"),

    ("horn_functional",
     r"Horn installed and functional\?\s*([^\n]+)"),

    ("brake_lights_functional",
     r"Brake lights & other lights functional\?\s*([^\n]+)"),

    ("vehicle_modified",
     r"Whether Vehicle Modified\s*([^\n]+)"),

    # =========================================================
    # ROAD DETAILS
    # =========================================================

    ("area_type", r"Area Type\s*([^\n]+)"),

    ("road_classification",
     r"Road Classification\s*([^\n]+)"),

    ("road_owning_agency",
     r"Road Owning Agency\s*([^\n]+)"),

    ("road_name",
     r"Road Name / Street Name\s*([^\n]+)"),

    ("road_surface_type",
     r"Type of Road Surface\s*([^\n]+)"),

    ("surface_condition",
     r"Surface Condition\s*([^\n]+)"),

    ("type_of_carriageway",
     r"Type of Carriageway\s*([^\n]+)"),

    ("road_width",
     r"Road Width \(in metre\)\s*([^\n]+)"),

    ("accident_location",
     r"Accident Location\s*([^\n]+)"),

    ("road_chainage",
     r"Road Chainage \(in metre\)\s*([^\n]+)"),

    ("horizontal_curve",
     r"Horizontal Curve\s*([^\n]+)"),

    ("vertical_curve",
     r"Vertical Curve\s*([^\n]+)"),

    ("junction_type",
     r"Junction Type\s*([^\n]+)"),

    ("junction_control",
     r"Junction Control\s*([^\n]+)"),

    ("speed_limit",
     r"Speed Limit \(in KMPH\)\s*([^\n]+)"),

    ("road_margins",
     r"Road Margins\s*([^\n]+)"),

    ("terrain_type",
     r"Type of Terrain\s*([^\n]+)"),

    ("surface_gradient",
     r"Type of Surface Gradient\s*([^\n]+)"),

    ("physical_divider",
     r"Physical Divider / Barrier Type\s*([^\n]+)"),

    ("median_type",
     r"Type of Median\s*([^\n]+)"),

    ("pedestrian_infrastructure",
     r"Pedestrian Infrastructure\s*([^\n]+)"),

    ("ongoing_road_work",
     r"Ongoing Road Work\s*([^\n]+)"),

    ("road_markings",
     r"Road Markings\s*([^\n]+)"),

    ("road_sign_board",
     r"Road Sign Board\s*([^\n]+)"),

    ("type_of_structure",
     r"Type of Structure\s*([^\n]+)"),

    ("type_of_road_surface",
     r"Type of Road Surface\s*([^\n]+)")
]

def extract_text_from_pdf(pdf_path: str) -> str:
    """Extract all text from PDF using pdfplumber."""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text

def clean_value(value):
    if value is None:
        return None

    value = str(value).strip()

    invalid_values = [
        "",
        "NA",
        "N/A",
        "na",
        "null",
        "None",
        "-"
    ]
    if value in invalid_values:
        return None
    return value

def extract_fields(text: str) -> dict:
    record = {}
    for field_name, pattern in FIELD_PATTERNS:
        match = re.search(
            pattern,
            text,
            re.IGNORECASE | re.DOTALL
        )
        if match:
            value = match.group(1)
            value = re.sub(r"\s+", " ", value).strip()
            value = clean_value(value)
            record[field_name] = value
        else:
            record[field_name] = None

    return record

EXPECTED_COLUMNS = [field[0] for field in FIELD_PATTERNS]

def validate_schema(df):
    missing_cols = []
    for col in EXPECTED_COLUMNS:
        if col not in df.columns:
            missing_cols.append(col)
    if missing_cols:
        print("Missing Columns:")
        print(missing_cols)

    return df

def remove_duplicate_accidents(df):
    before = len(df)
    df = df.drop_duplicates(
        subset=["accident_id"],
        keep="first"
    )
    after = len(df)
    print(f"Duplicates Removed: {before - after}")

    return df

def process_multiple_pdfs(pdf_paths: list):
    records = []
    for path in pdf_paths:
        print(f"Processing: {os.path.basename(path)}")
        try:
            text = extract_text_from_pdf(path)
            record = extract_fields(text)
            # Place source_pdf as the first column for easy identification.
            record = {"source_pdf": os.path.basename(path), **record}
            records.append(record)
        except Exception as e:
            print(f"Error processing {path}")
            print(e)

    df = pd.DataFrame(records)
    numeric_columns = [
    "visibility",
    "approximate_damage_value",
    "number_of_vehicles_involved",
    "number_of_persons_involved",
    "number_of_animals_involved",
    "driver_killed",
    "driver_grievous_injury",
    "driver_minor_injury",
    "driver_no_injury",
    "passenger_killed",
    "passenger_grievous_injury",
    "passenger_minor_injury",
    "passenger_no_injury",
    "pedestrian_killed",
    "pedestrian_grievous_injury",
    "pedestrian_minor_injury",
    "pedestrian_no_injury",
    "total_killed",
    "total_grievous_injury",
    "total_minor_injury",
    "total_no_injury",
    "total_persons_involved",
    "age",
    "vehicle_laden_weight",
    "vehicle_unladen_weight",
    "seating_capacity"
]

    for col in numeric_columns:

        if col in df.columns:

            df[col] = pd.to_numeric(
                df[col],
                errors="coerce"
            )
    df = validate_schema(df)
    df = remove_duplicate_accidents(df)

    return df

def apply_excel_formatting(output_path: str):
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "Crash Data"

    # HEADER STYLE
    header_fill = PatternFill(
        fill_type="solid",
        start_color="D9D9D9",
        end_color="D9D9D9"
    )
    header_font = Font(
        bold=True,
        color="000000",
        name="Arial",
        size=10
    )
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    # NORMAL CELL STYLE
    normal_fill = PatternFill(
        fill_type="solid",
        start_color="FFFFFF",   # White
        end_color="FFFFFF"
    )
    normal_font = Font(
        name="Arial",
        size=9
    )
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    max_row = ws.max_row
    max_col = ws.max_column

    # FORMAT HEADER
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # FORMAT DATA ROWS
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = normal_fill
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=False
            )

    # AUTO WIDTH
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        max_length = 10
        for row in range(1, max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_length = max(
                    max_length,
                    min(len(str(value)), 40)
                )
        ws.column_dimensions[column_letter].width = max_length + 2
    # Freeze Header
    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"Formatting applied → {output_path}")


def main():
    input_folder = "./pdfs"
    output_excel = "./output/crash_reports.xlsx"
    pdf_files = glob.glob(os.path.join(input_folder, "*.pdf"))

    if not pdf_files:
        print("No PDFs found.")
        return

    print("=" * 60)
    print("Crash Report PDF to Excel Extractor")
    print("=" * 60)
    print(f"\nFound {len(pdf_files)} PDF files\n")

    df = process_multiple_pdfs(pdf_files)
    df.to_excel(
        output_excel,
        index=False,
        engine="openpyxl"
    )
    apply_excel_formatting(output_excel)
    print("\nExtraction Completed")
    print(f"Excel saved at: {output_excel}")

if __name__ == "__main__":
    main()